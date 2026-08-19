from datetime import date, datetime, time, timedelta
from io import BytesIO
import re
import unicodedata

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import and_, or_
from sqlalchemy.orm import Session, joinedload

from ..models import AttendanceSession, DahWebhookEvent, Employee, EmployeeShiftSchedule
from ..timeutils import utc_now
from .audit_service import record_audit

DEFAULT_EARLY_CHECKIN_MINUTES = 15
EXCEL_DAY_LABELS = ("T2", "T3", "T4", "T5", "T6", "T7", "CN")


def _name_key(value: str | None):
    normalized = unicodedata.normalize("NFD", str(value or "").strip().lower())
    no_marks = "".join(char for char in normalized if unicodedata.category(char) != "Mn")
    return " ".join(no_marks.replace("đ", "d").split())


def _cell_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def _parse_excel_week_start(rows):
    title = " ".join(_cell_text(cell) for row in rows[:4] for cell in row)
    match = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", title)
    if not match:
        return date.today() - timedelta(days=date.today().weekday())
    day, month, year = (int(part) for part in match.groups())
    return date(year, month, day)


def preview_employee_shift_excel(file_bytes: bytes, filename: str = ""):
    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise HTTPException(422, "Không đọc được file Excel .xlsx.") from exc
    sheet = workbook[workbook.sheetnames[0]]
    rows = [[cell.value for cell in row] for row in sheet.iter_rows()]
    header_index = None
    name_col = None
    position_col = None
    for index, row in enumerate(rows):
        normalized = [_name_key(_cell_text(cell)) for cell in row]
        possible_name = next((col for col, text in enumerate(normalized) if "ho va ten" in text), None)
        possible_position = next((col for col, text in enumerate(normalized) if "vi tri" in text), None)
        if possible_name is not None and possible_position is not None:
            header_index = index
            name_col = possible_name
            position_col = possible_position
            break
    if header_index is None:
        raise HTTPException(422, "Không tìm thấy dòng HỌ VÀ TÊN / VỊ TRÍ trong file.")
    week_start = _parse_excel_week_start(rows)
    first_day_col = max(name_col, position_col) + 1
    days = [
        {
            "label": label,
            "workDate": (week_start + timedelta(days=index)).isoformat(),
            "column": first_day_col + index,
        }
        for index, label in enumerate(EXCEL_DAY_LABELS)
    ]
    preview_rows = []
    for row_index, row in enumerate(rows[header_index + 2:]):
        employee_name = _cell_text(row[name_col] if name_col < len(row) else "")
        position = _cell_text(row[position_col] if position_col < len(row) else "")
        cells = [
            _cell_text(row[day["column"]] if day["column"] < len(row) else "")
            for day in days
        ]
        if not employee_name or not any(cells):
            continue
        preview_rows.append({
            "id": f"{row_index}-{employee_name}",
            "employeeName": employee_name,
            "position": position,
            "cells": cells,
        })
    if not preview_rows:
        raise HTTPException(422, "File không có dòng lịch nhân viên để nhập.")
    return {
        "sourceName": filename,
        "weekStart": week_start.isoformat(),
        "days": [{key: value for key, value in day.items() if key != "column"} for day in days],
        "rows": preview_rows,
    }


def _as_date(value):
    if isinstance(value, date):
        return value
    if not value:
        return None
    try:
        return date.fromisoformat(str(value))
    except (TypeError, ValueError) as exc:
        raise HTTPException(422, "Ngày làm việc không hợp lệ. Vui lòng dùng định dạng YYYY-MM-DD.") from exc


def _as_time(value):
    if isinstance(value, time):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return time.fromisoformat(text)
    except (TypeError, ValueError) as exc:
        raise HTTPException(422, "Giờ làm việc không hợp lệ. Vui lòng dùng định dạng HH:MM.") from exc


def _time_text(value: datetime | None):
    return value.strftime("%H:%M") if value else None


def _shift_data(row: EmployeeShiftSchedule):
    return {
        "id": row.id,
        "employeeId": row.employee_id,
        "workDate": row.work_date.isoformat(),
        "startsAt": row.starts_at.isoformat(),
        "endsAt": row.ends_at.isoformat(),
        "startTime": _time_text(row.starts_at),
        "endTime": _time_text(row.ends_at),
        "status": row.status,
        "note": row.note,
    }


def list_employee_shifts(db: Session, employee_id: int, date_from: str = "", date_to: str = ""):
    employee = db.get(Employee, employee_id)
    if not employee:
        raise HTTPException(404, "Không tìm thấy nhân viên.")
    start = _as_date(date_from) or date.today()
    end = _as_date(date_to) or start + timedelta(days=6)
    if end < start:
        raise HTTPException(422, "Ngày kết thúc phải sau ngày bắt đầu.")
    rows = (
        db.query(EmployeeShiftSchedule)
        .filter(
            EmployeeShiftSchedule.employee_id == employee_id,
            EmployeeShiftSchedule.work_date >= start,
            EmployeeShiftSchedule.work_date <= end,
            EmployeeShiftSchedule.status != "deleted",
        )
        .order_by(EmployeeShiftSchedule.work_date.asc(), EmployeeShiftSchedule.starts_at.asc(), EmployeeShiftSchedule.id.asc())
        .all()
    )
    return {"employeeId": employee_id, "items": [_shift_data(row) for row in rows]}


def list_employee_shifts_week(db: Session, date_from: str = "", date_to: str = ""):
    start = _as_date(date_from) or date.today()
    end = _as_date(date_to) or start + timedelta(days=6)
    if end < start:
        raise HTTPException(422, "Ngày kết thúc phải sau ngày bắt đầu.")
    rows = (
        db.query(EmployeeShiftSchedule)
        .options(joinedload(EmployeeShiftSchedule.employee).joinedload(Employee.person))
        .filter(
            EmployeeShiftSchedule.work_date >= start,
            EmployeeShiftSchedule.work_date <= end,
            EmployeeShiftSchedule.status != "deleted",
        )
        .order_by(
            EmployeeShiftSchedule.work_date.asc(),
            EmployeeShiftSchedule.employee_id.asc(),
            EmployeeShiftSchedule.starts_at.asc(),
            EmployeeShiftSchedule.id.asc(),
        )
        .all()
    )
    return {"dateFrom": start.isoformat(), "dateTo": end.isoformat(), "items": [_shift_data(row) for row in rows]}


def _validate_shift_payload(payload: dict):
    work_date = _as_date(payload.get("workDate"))
    starts = _as_time(payload.get("startTime"))
    ends = _as_time(payload.get("endTime"))
    if not work_date:
        raise HTTPException(422, "Ngày làm là bắt buộc.")
    if not starts or not ends:
        raise HTTPException(422, "Giờ bắt đầu và kết thúc là bắt buộc.")
    starts_at = datetime.combine(work_date, starts)
    ends_at = datetime.combine(work_date, ends)
    if ends_at <= starts_at:
        raise HTTPException(422, "Giờ kết thúc phải sau giờ bắt đầu.")
    note = str(payload.get("note") or "").strip()[:255] or None
    return work_date, starts_at, ends_at, note


def _ensure_no_overlap(db: Session, employee_id: int, work_date: date, starts_at: datetime, ends_at: datetime, exclude_id: int | None = None):
    query = db.query(EmployeeShiftSchedule).filter(
        EmployeeShiftSchedule.employee_id == employee_id,
        EmployeeShiftSchedule.work_date == work_date,
        EmployeeShiftSchedule.status == "active",
        EmployeeShiftSchedule.starts_at < ends_at,
        EmployeeShiftSchedule.ends_at > starts_at,
    )
    if exclude_id:
        query = query.filter(EmployeeShiftSchedule.id != exclude_id)
    if query.first():
        raise HTTPException(409, "Ca làm bị trùng với ca đang có của nhân viên.")


def create_employee_shift(db: Session, employee_id: int, payload: dict, actor=None):
    employee = db.get(Employee, employee_id)
    if not employee or employee.status != "active":
        raise HTTPException(404, "Không tìm thấy nhân viên đang hoạt động.")
    work_date, starts_at, ends_at, note = _validate_shift_payload(payload)
    _ensure_no_overlap(db, employee_id, work_date, starts_at, ends_at)
    row = EmployeeShiftSchedule(
        employee_id=employee_id,
        work_date=work_date,
        starts_at=starts_at,
        ends_at=ends_at,
        status="active",
        note=note,
    )
    db.add(row)
    db.flush()
    record_audit(
        db,
        actor,
        "create",
        "employee_shift_schedule",
        row.id,
        f"Thêm ca làm {employee.person.display_name} {starts_at.strftime('%d/%m %H:%M')}-{ends_at.strftime('%H:%M')}",
        details={"employeeId": employee_id, "workDate": work_date.isoformat()},
    )
    rebuild_employee_attendance_for_day(db, employee_id, work_date)
    db.commit()
    db.refresh(row)
    return _shift_data(row)


def create_employee_shifts_bulk(db: Session, employee_id: int, payload: dict, actor=None):
    employee = db.get(Employee, employee_id)
    if not employee or employee.status != "active":
        raise HTTPException(404, "Không tìm thấy nhân viên đang hoạt động.")
    week_start = _as_date(payload.get("weekStart"))
    weekdays = payload.get("weekdays") or []
    starts = _as_time(payload.get("startTime"))
    ends = _as_time(payload.get("endTime"))
    note = str(payload.get("note") or "").strip()[:255] or None
    if not week_start:
        raise HTTPException(422, "Ngày bắt đầu tuần là bắt buộc.")
    try:
        offsets = sorted({int(day) for day in weekdays})
    except (TypeError, ValueError):
        raise HTTPException(422, "Danh sách ngày trong tuần không hợp lệ.")
    if not offsets or any(day < 0 or day > 6 for day in offsets):
        raise HTTPException(422, "Chọn ít nhất một ngày từ Thứ 2 đến Chủ nhật.")
    if not starts or not ends:
        raise HTTPException(422, "Giờ bắt đầu và kết thúc là bắt buộc.")

    created = []
    skipped = []
    affected_dates = set()
    for offset in offsets:
        work_date = week_start + timedelta(days=offset)
        starts_at = datetime.combine(work_date, starts)
        ends_at = datetime.combine(work_date, ends)
        if ends_at <= starts_at:
            raise HTTPException(422, "Giờ kết thúc phải sau giờ bắt đầu.")
        try:
            _ensure_no_overlap(db, employee_id, work_date, starts_at, ends_at)
        except HTTPException as exc:
            if exc.status_code == 409:
                skipped.append({
                    "workDate": work_date.isoformat(),
                    "reason": "Ca bị trùng với lịch đang có.",
                })
                continue
            raise
        row = EmployeeShiftSchedule(
            employee_id=employee_id,
            work_date=work_date,
            starts_at=starts_at,
            ends_at=ends_at,
            status="active",
            note=note,
        )
        db.add(row)
        db.flush()
        created.append(row)
        affected_dates.add(work_date)

    if not created and skipped:
        raise HTTPException(409, "Tất cả ca đã chọn đều bị trùng lịch hiện có.")
    for work_date in affected_dates:
        rebuild_employee_attendance_for_day(db, employee_id, work_date)
    record_audit(
        db,
        actor,
        "create",
        "employee_shift_schedule",
        None,
        f"Tạo nhanh {len(created)} ca làm cho {employee.person.display_name}",
        details={
            "employeeId": employee_id,
            "weekStart": week_start.isoformat(),
            "weekdays": offsets,
            "created": len(created),
            "skipped": len(skipped),
        },
    )
    db.commit()
    return {
        "created": len(created),
        "skipped": skipped,
        "items": [_shift_data(row) for row in created],
    }


def import_employee_shifts(db: Session, payload: dict, actor=None):
    rows = payload.get("rows") or []
    if not isinstance(rows, list) or not rows:
        raise HTTPException(422, "Không có dữ liệu lịch để nhập.")
    employees = (
        db.query(Employee)
        .options(joinedload(Employee.person))
        .filter(Employee.status == "active")
        .all()
    )
    by_name = {}
    duplicates = set()
    for employee in employees:
        key = _name_key(employee.person.display_name if employee.person else "")
        if not key:
            continue
        if key in by_name:
            duplicates.add(key)
        else:
            by_name[key] = employee

    created = []
    skipped = []
    unmatched = []
    affected = set()
    for row_index, row in enumerate(rows):
        employee_id = row.get("employeeId")
        employee = db.get(Employee, employee_id) if employee_id else None
        employee_name = str(row.get("employeeName") or "").strip()
        if not employee:
            key = _name_key(employee_name)
            if key in duplicates:
                unmatched.append({"row": row_index, "employeeName": employee_name, "reason": "Trùng tên nhân viên trong hệ thống."})
                continue
            employee = by_name.get(key)
        if not employee or employee.status != "active":
            unmatched.append({"row": row_index, "employeeName": employee_name, "reason": "Không tìm thấy nhân viên đang hoạt động."})
            continue
        for shift_index, shift in enumerate(row.get("shifts") or []):
            try:
                work_date, starts_at, ends_at, note = _validate_shift_payload({
                    "workDate": shift.get("workDate"),
                    "startTime": shift.get("startTime"),
                    "endTime": shift.get("endTime"),
                    "note": shift.get("note") or row.get("position") or None,
                })
            except HTTPException as exc:
                skipped.append({
                    "row": row_index,
                    "shift": shift_index,
                    "employeeName": employee_name,
                    "reason": exc.detail,
                })
                continue
            try:
                _ensure_no_overlap(db, employee.id, work_date, starts_at, ends_at)
            except HTTPException as exc:
                skipped.append({
                    "row": row_index,
                    "shift": shift_index,
                    "employeeName": employee.person.display_name if employee.person else employee_name,
                    "workDate": work_date.isoformat(),
                    "reason": exc.detail,
                })
                continue
            schedule = EmployeeShiftSchedule(
                employee_id=employee.id,
                work_date=work_date,
                starts_at=starts_at,
                ends_at=ends_at,
                status="active",
                note=note,
            )
            db.add(schedule)
            db.flush()
            created.append(schedule)
            affected.add((employee.id, work_date))
    for employee_id, work_date in affected:
        rebuild_employee_attendance_for_day(db, employee_id, work_date)
    record_audit(
        db,
        actor,
        "import",
        "employee_shift_schedule",
        None,
        f"Nhập lịch Excel: {len(created)} ca",
        details={
            "created": len(created),
            "skipped": len(skipped),
            "unmatched": len(unmatched),
            "source": payload.get("sourceName"),
        },
    )
    db.commit()
    return {
        "created": len(created),
        "skipped": skipped,
        "unmatched": unmatched,
        "items": [_shift_data(row) for row in created],
    }


def replace_employee_shifts_week(db: Session, payload: dict, actor=None):
    week_start = _as_date(payload.get("weekStart"))
    rows = payload.get("rows") or []
    if not week_start:
        raise HTTPException(422, "Ngày bắt đầu tuần là bắt buộc.")
    if not isinstance(rows, list):
        raise HTTPException(422, "Dữ liệu lịch tuần không hợp lệ.")
    week_end = week_start + timedelta(days=6)
    employee_ids = []
    for row in rows:
        try:
            employee_ids.append(int(row.get("employeeId")))
        except (TypeError, ValueError):
            raise HTTPException(422, "Dòng lịch thiếu nhân viên.")
    employee_ids = sorted(set(employee_ids))
    if not employee_ids:
        raise HTTPException(422, "Không có nhân viên để lưu lịch.")

    employees = {
        employee.id: employee
        for employee in db.query(Employee)
        .options(joinedload(Employee.person))
        .filter(Employee.id.in_(employee_ids), Employee.status == "active")
        .all()
    }
    missing = [employee_id for employee_id in employee_ids if employee_id not in employees]
    if missing:
        raise HTTPException(404, "Có nhân viên không tồn tại hoặc đã ngưng hoạt động.")

    existing = (
        db.query(EmployeeShiftSchedule)
        .filter(
            EmployeeShiftSchedule.employee_id.in_(employee_ids),
            EmployeeShiftSchedule.work_date >= week_start,
            EmployeeShiftSchedule.work_date <= week_end,
            EmployeeShiftSchedule.status == "active",
        )
        .all()
    )
    deleted = 0
    for row in existing:
        row.status = "deleted"
        row.updated_at = utc_now()
        deleted += 1

    created = []
    affected = {(employee_id, week_start + timedelta(days=offset)) for employee_id in employee_ids for offset in range(7)}
    seen = set()
    for row_index, row in enumerate(rows):
        employee_id = int(row.get("employeeId"))
        employee = employees[employee_id]
        for shift_index, shift in enumerate(row.get("shifts") or []):
            work_date, starts_at, ends_at, note = _validate_shift_payload({
                "workDate": shift.get("workDate"),
                "startTime": shift.get("startTime"),
                "endTime": shift.get("endTime"),
                "note": shift.get("note") or None,
            })
            if work_date < week_start or work_date > week_end:
                raise HTTPException(422, "Ca làm nằm ngoài tuần đang lưu.")
            duplicate_key = (employee_id, starts_at, ends_at)
            if duplicate_key in seen:
                raise HTTPException(409, f"Ca bị trùng trong dữ liệu nhập ở dòng {row_index + 1}.")
            seen.add(duplicate_key)
            for other in created:
                if other.employee_id == employee_id and other.work_date == work_date and other.starts_at < ends_at and other.ends_at > starts_at:
                    raise HTTPException(409, f"Ca bị trùng trong dữ liệu nhập ở dòng {row_index + 1}.")
            schedule = EmployeeShiftSchedule(
                employee_id=employee_id,
                work_date=work_date,
                starts_at=starts_at,
                ends_at=ends_at,
                status="active",
                note=note,
            )
            db.add(schedule)
            db.flush()
            created.append(schedule)
            affected.add((employee_id, work_date))

    for employee_id, work_date in affected:
        rebuild_employee_attendance_for_day(db, employee_id, work_date)
    record_audit(
        db,
        actor,
        "replace",
        "employee_shift_schedule",
        None,
        f"Lưu lịch tuần {week_start.isoformat()} - {week_end.isoformat()}: {len(created)} ca",
        details={
            "weekStart": week_start.isoformat(),
            "weekEnd": week_end.isoformat(),
            "employeeCount": len(employee_ids),
            "created": len(created),
            "deleted": deleted,
        },
    )
    db.commit()
    return {"created": len(created), "deleted": deleted, "items": [_shift_data(row) for row in created]}


def update_employee_shift(db: Session, shift_id: int, payload: dict, actor=None):
    row = (
        db.query(EmployeeShiftSchedule)
        .options(joinedload(EmployeeShiftSchedule.employee).joinedload(Employee.person))
        .filter(EmployeeShiftSchedule.id == shift_id, EmployeeShiftSchedule.status != "deleted")
        .first()
    )
    if not row:
        raise HTTPException(404, "Không tìm thấy ca làm.")
    old_date = row.work_date
    work_date, starts_at, ends_at, note = _validate_shift_payload({
        "workDate": payload.get("workDate", row.work_date.isoformat()),
        "startTime": payload.get("startTime", _time_text(row.starts_at)),
        "endTime": payload.get("endTime", _time_text(row.ends_at)),
        "note": payload.get("note", row.note),
    })
    _ensure_no_overlap(db, row.employee_id, work_date, starts_at, ends_at, exclude_id=row.id)
    row.work_date = work_date
    row.starts_at = starts_at
    row.ends_at = ends_at
    row.note = note
    row.updated_at = utc_now()
    record_audit(
        db,
        actor,
        "update",
        "employee_shift_schedule",
        row.id,
        f"Cập nhật ca làm {row.employee.person.display_name if row.employee and row.employee.person else row.employee_id}",
        details={"fields": list(payload.keys()), "oldWorkDate": old_date.isoformat(), "workDate": work_date.isoformat()},
    )
    rebuild_employee_attendance_for_day(db, row.employee_id, old_date)
    if work_date != old_date:
        rebuild_employee_attendance_for_day(db, row.employee_id, work_date)
    db.commit()
    db.refresh(row)
    return _shift_data(row)


def delete_employee_shift(db: Session, shift_id: int, actor=None):
    row = (
        db.query(EmployeeShiftSchedule)
        .options(joinedload(EmployeeShiftSchedule.employee).joinedload(Employee.person))
        .filter(EmployeeShiftSchedule.id == shift_id, EmployeeShiftSchedule.status != "deleted")
        .first()
    )
    if not row:
        raise HTTPException(404, "Không tìm thấy ca làm.")
    employee_id = row.employee_id
    work_date = row.work_date
    row.status = "deleted"
    row.updated_at = utc_now()
    record_audit(
        db,
        actor,
        "delete",
        "employee_shift_schedule",
        row.id,
        f"Xóa ca làm {row.employee.person.display_name if row.employee and row.employee.person else employee_id}",
        details={"employeeId": employee_id, "workDate": work_date.isoformat()},
    )
    rebuild_employee_attendance_for_day(db, employee_id, work_date)
    db.commit()
    return {"deleted": True, "id": shift_id}


def _day_bounds(work_date: date):
    start = datetime.combine(work_date, datetime.min.time())
    return start, start + timedelta(days=1)


def _employee_scan_events(db: Session, employee_id: int, work_date: date, extra_event_time: datetime | None = None):
    start, end = _day_bounds(work_date)
    rows = (
        db.query(DahWebhookEvent)
        .filter(
            DahWebhookEvent.employee_id == employee_id,
            DahWebhookEvent.event_time >= start,
            DahWebhookEvent.event_time < end,
            DahWebhookEvent.verify_status == 1,
            DahWebhookEvent.status == "processed",
            DahWebhookEvent.action.in_(("checkin", "checkout", "mixed", "employee_shift_sync", "local_sync")),
        )
        .order_by(DahWebhookEvent.event_time.asc(), DahWebhookEvent.id.asc())
        .all()
    )
    times = [row.event_time for row in rows if row.event_time]
    if extra_event_time and start <= extra_event_time < end:
        times.append(extra_event_time)
    return sorted(set(times))


def _active_shifts(db: Session, employee_id: int, work_date: date):
    return (
        db.query(EmployeeShiftSchedule)
        .filter(
            EmployeeShiftSchedule.employee_id == employee_id,
            EmployeeShiftSchedule.work_date == work_date,
            EmployeeShiftSchedule.status == "active",
        )
        .order_by(EmployeeShiftSchedule.starts_at.asc(), EmployeeShiftSchedule.id.asc())
        .all()
    )


def _boundary_cutoff(previous: EmployeeShiftSchedule, current: EmployeeShiftSchedule):
    gap = max((current.starts_at - previous.ends_at).total_seconds(), 0)
    default_window = DEFAULT_EARLY_CHECKIN_MINUTES * 60
    early_seconds = gap / 4 if gap <= default_window else default_window
    return current.starts_at - timedelta(seconds=early_seconds)


def _assign_events_to_shifts(shifts: list[EmployeeShiftSchedule], events: list[datetime]):
    if not events:
        return []
    if not shifts:
        return [{
            "shift": None,
            "events": events,
            "scheduled_start": None,
            "scheduled_end": None,
        }]
    if len(shifts) == 1:
        return [{
            "shift": shifts[0],
            "events": events,
            "scheduled_start": shifts[0].starts_at,
            "scheduled_end": shifts[0].ends_at,
        }]

    cutoffs = [_boundary_cutoff(shifts[index - 1], shifts[index]) for index in range(1, len(shifts))]
    buckets = [[] for _ in shifts]
    for event_time in events:
        bucket_index = 0
        for index, cutoff in enumerate(cutoffs):
            if event_time >= cutoff:
                bucket_index = index + 1
            else:
                break
        buckets[bucket_index].append(event_time)
    return [
        {
            "shift": shift,
            "events": bucket,
            "scheduled_start": shift.starts_at,
            "scheduled_end": shift.ends_at,
        }
        for shift, bucket in zip(shifts, buckets)
    ]


def rebuild_employee_attendance_for_day(
    db: Session,
    employee_id: int,
    work_date: date,
    extra_event_time: datetime | None = None,
):
    events = _employee_scan_events(db, employee_id, work_date, extra_event_time=extra_event_time)
    shifts = _active_shifts(db, employee_id, work_date)
    start, end = _day_bounds(work_date)
    existing = (
        db.query(AttendanceSession)
        .filter(
            AttendanceSession.employee_id == employee_id,
            or_(
                and_(AttendanceSession.checked_in_at >= start, AttendanceSession.checked_in_at < end),
                and_(AttendanceSession.scheduled_start_at >= start, AttendanceSession.scheduled_start_at < end),
            ),
            AttendanceSession.source == "dah",
        )
        .all()
    )
    existing_ids = [row.id for row in existing]
    if existing_ids:
        db.query(DahWebhookEvent).filter(
            DahWebhookEvent.attendance_session_id.in_(existing_ids)
        ).update({DahWebhookEvent.attendance_session_id: None}, synchronize_session=False)
    for row in existing:
        db.delete(row)
    db.flush()

    sessions = []
    for assignment in _assign_events_to_shifts(shifts, events):
        bucket = assignment["events"]
        if not bucket:
            continue
        checked_in_at = bucket[0]
        checked_out_at = bucket[-1] if len(bucket) > 1 else None
        shift = assignment["shift"]
        note_parts = ["DAH theo ca"]
        if shift:
            note_parts.append(f"{shift.starts_at.strftime('%H:%M')}-{shift.ends_at.strftime('%H:%M')}")
        if len(bucket) == 1:
            note_parts.append("thiếu check-out")
        session = AttendanceSession(
            employee_id=employee_id,
            employee_shift_schedule_id=shift.id if shift else None,
            scheduled_start_at=assignment["scheduled_start"],
            scheduled_end_at=assignment["scheduled_end"],
            checked_in_at=checked_in_at,
            checked_out_at=checked_out_at,
            source="dah",
            result="allowed",
            status="closed" if checked_out_at else "open",
            note=" · ".join(note_parts)[:255],
        )
        db.add(session)
        db.flush()
        sessions.append(session)
    return sessions


def sync_employee_scan(db: Session, employee: Employee, event_time: datetime, device=None):
    if employee.status != "active":
        return {"status": "denied", "action": "denied", "note": "Nhân viên không ở trạng thái hoạt động.", "session_id": None}
    sessions = rebuild_employee_attendance_for_day(db, employee.id, event_time.date(), extra_event_time=event_time)
    target = None
    for session in sessions:
        if session.checked_in_at == event_time or session.checked_out_at == event_time:
            target = session
            break
    if not target and sessions:
        target = sessions[-1]
    if not target:
        return {"status": "denied", "action": "denied", "note": "Không thể ghi nhận chấm công nhân viên.", "session_id": None}
    action = "checkout" if target.checked_out_at == event_time and target.checked_in_at != event_time else "checkin"
    return {
        "status": "processed",
        "action": action,
        "note": "Đã đồng bộ chấm công nhân viên theo ca.",
        "session_id": target.id,
    }
