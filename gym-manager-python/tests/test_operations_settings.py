import os
from pathlib import Path
import tempfile
from datetime import date, datetime, timedelta
from io import BytesIO

from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

os.environ.setdefault("GYM_ENV", "test")
os.environ.setdefault(
    "GYM_DATABASE_PATH",
    str(Path(tempfile.mkdtemp(prefix="pulsefit-operations-tests-")) / "bootstrap.sqlite3"),
)


def make_session(tmp_path):
    from server.database import Base

    engine = create_engine(f"sqlite:///{tmp_path / 'settings.sqlite3'}")
    Base.metadata.create_all(bind=engine)
    return sessionmaker(bind=engine)()


def active_titles(db):
    from server.models import EmployeeJobTitle

    return {
        row.name
        for row in db.query(EmployeeJobTitle)
        .filter(EmployeeJobTitle.is_active == True)
        .all()
    }


def test_renaming_default_job_title_does_not_recreate_old_name(tmp_path):
    from server.models import EmployeeJobTitle
    from server.services.operations_service import (
        ensure_employee_job_titles,
        update_job_title,
    )

    db = make_session(tmp_path)
    try:
        ensure_employee_job_titles(db)
        db.commit()
        sale = db.query(EmployeeJobTitle).filter(EmployeeJobTitle.name == "Sale").one()

        update_job_title(db, sale.id, {"name": "sale f", "renameEmployees": True})
        ensure_employee_job_titles(db)
        db.commit()

        names = active_titles(db)
        assert "sale f" in names
        assert "Sale" not in names
    finally:
        db.close()


def test_deleting_default_job_title_does_not_recreate_it(tmp_path):
    from server.models import EmployeeJobTitle
    from server.services.operations_service import (
        delete_job_title,
        ensure_employee_job_titles,
    )

    db = make_session(tmp_path)
    try:
        ensure_employee_job_titles(db)
        db.commit()
        marketing = db.query(EmployeeJobTitle).filter(EmployeeJobTitle.name == "Marketing").one()

        delete_job_title(db, marketing.id)
        ensure_employee_job_titles(db)
        db.commit()

        assert "Marketing" not in active_titles(db)
    finally:
        db.close()


def test_settings_returns_single_dah1017_with_heartbeat_status(tmp_path):
    from server.models import Device
    from server.services.operations_service import settings
    from server.timeutils import utc_now

    db = make_session(tmp_path)
    try:
        db.add(Device(code="OLD-GATE", name="Old Gate", model="Other", status="online"))
        db.add(Device(code="DAH-2470802", name="DAH 2470802", model="DAH1017", status="online", last_heartbeat_at=utc_now() - timedelta(seconds=120)))
        db.commit()

        data = settings(db)
        assert len(data["devices"]) == 1
        assert data["devices"][0]["name"] == "DAH1017"
        assert data["devices"][0]["status"] == "offline"

        db.query(Device).filter(Device.code == "DAH-2470802").one().last_heartbeat_at = utc_now()
        db.commit()
        assert settings(db)["devices"][0]["status"] == "online"
    finally:
        db.close()


def test_trainers_show_pt_client_status_counts_only_for_pt_roles(tmp_path):
    from server.models import Customer, Employee, Person, PtEnrollment, PtEnrollmentCoach
    from server.services.operations_service import ensure_employee_job_titles, list_trainers

    db = make_session(tmp_path)
    try:
        coach_person = Person(display_name="Coach One", phone="0900000001", status="active")
        sale_person = Person(display_name="Sale One", phone="0900000002", status="active")
        active_member_person = Person(display_name="Active PT Member", phone="0900000003", status="active")
        expired_member_person = Person(display_name="Expired PT Member", phone="0900000004", status="active")
        db.add_all([coach_person, sale_person, active_member_person, expired_member_person])
        db.flush()

        coach = Employee(person_id=coach_person.id, employee_code="EMP-00001", job_title="Coach", status="active")
        sale = Employee(person_id=sale_person.id, employee_code="EMP-00002", job_title="Sale", status="active")
        active_member = Customer(person_id=active_member_person.id, customer_code="CUS0000001", status="active")
        expired_member = Customer(person_id=expired_member_person.id, customer_code="CUS0000002", status="active")
        db.add_all([coach, sale, active_member, expired_member])
        db.flush()

        active = PtEnrollment(
            customer_id=active_member.id,
            group_type="1:1",
            starts_at=date(2026, 8, 1),
            expires_at=date(2026, 9, 1),
            total_sessions=12,
            remaining_sessions=12,
            status="active",
        )
        expired = PtEnrollment(
            customer_id=expired_member.id,
            group_type="1:1",
            starts_at=date(2026, 6, 1),
            expires_at=date(2026, 7, 1),
            total_sessions=12,
            remaining_sessions=0,
            status="active",
        )
        db.add_all([active, expired])
        db.flush()
        db.add_all([
            PtEnrollmentCoach(enrollment_id=active.id, coach_id=coach.id),
            PtEnrollmentCoach(enrollment_id=expired.id, coach_id=coach.id),
        ])
        ensure_employee_job_titles(db)
        db.commit()

        rows = {row["name"]: row for row in list_trainers(db, q="", title="all", page=1, page_size=20)["items"]}

        assert rows["Coach One"]["isPtRole"] is True
        assert rows["Coach One"]["registeredPtClients"] == 2
        assert rows["Coach One"]["activePtClients"] == 1
        assert rows["Coach One"]["expiredPtClients"] == 1
        assert "ptSessions" not in rows["Coach One"]
        assert rows["Sale One"]["isPtRole"] is False
        assert rows["Sale One"]["registeredPtClients"] is None
        assert rows["Sale One"]["activePtClients"] is None
        assert rows["Sale One"]["expiredPtClients"] is None
    finally:
        db.close()


def test_employee_attendance_returns_multiple_shifts_in_one_day(tmp_path):
    from server.models import AttendanceSession, Employee, Person
    from server.services.operations_service import employee_attendance

    db = make_session(tmp_path)
    try:
        person = Person(display_name="Coach Shift", phone="0900000010", status="active")
        db.add(person)
        db.flush()
        employee = Employee(person_id=person.id, employee_code="EMP-00010", job_title="Coach", status="active")
        db.add(employee)
        db.flush()
        db.add_all([
            AttendanceSession(
                employee_id=employee.id,
                checked_in_at=datetime(2026, 8, 12, 8, 0),
                checked_out_at=datetime(2026, 8, 12, 12, 0),
                source="dah",
                status="closed",
            ),
            AttendanceSession(
                employee_id=employee.id,
                checked_in_at=datetime(2026, 8, 12, 14, 0),
                checked_out_at=datetime(2026, 8, 12, 18, 30),
                source="dah",
                status="closed",
            ),
        ])
        db.commit()

        data = employee_attendance(db, "2026-08-12")

        assert [row["shiftNo"] for row in data["items"]] == [1, 2]
        assert [row["durationMinutes"] for row in data["items"]] == [240, 270]
        assert all(row["employeeId"] == employee.id for row in data["items"])
    finally:
        db.close()


def test_bulk_employee_shifts_create_week_and_skip_overlaps(tmp_path):
    from server.models import Employee, EmployeeShiftSchedule, Person
    from server.services.operations_service import create_employee_shifts_bulk

    db = make_session(tmp_path)
    try:
        person = Person(display_name="Bulk Shift", phone="0900000011", status="active")
        db.add(person)
        db.flush()
        employee = Employee(person_id=person.id, employee_code="EMP-00011", job_title="Sale", status="active")
        db.add(employee)
        db.flush()
        db.add(EmployeeShiftSchedule(
            employee_id=employee.id,
            work_date=date(2026, 8, 18),
            starts_at=datetime(2026, 8, 18, 8, 0),
            ends_at=datetime(2026, 8, 18, 12, 0),
        ))
        db.commit()

        result = create_employee_shifts_bulk(db, employee.id, {
            "weekStart": "2026-08-17",
            "weekdays": [0, 1, 2],
            "startTime": "08:00",
            "endTime": "12:00",
            "note": "Ca sáng",
        })

        rows = (
            db.query(EmployeeShiftSchedule)
            .filter(EmployeeShiftSchedule.employee_id == employee.id, EmployeeShiftSchedule.status == "active")
            .order_by(EmployeeShiftSchedule.work_date.asc())
            .all()
        )
        assert result["created"] == 2
        assert [row["workDate"] for row in result["skipped"]] == ["2026-08-18"]
        assert [row.work_date.isoformat() for row in rows] == ["2026-08-17", "2026-08-18", "2026-08-19"]
    finally:
        db.close()


def test_import_employee_shifts_matches_names_and_skips_overlaps(tmp_path):
    from server.models import Employee, EmployeeShiftSchedule, Person
    from server.services.operations_service import import_employee_shifts

    db = make_session(tmp_path)
    try:
        person = Person(display_name="Trần Nguyễn Khải Hoàn", phone="0900000012", status="active")
        db.add(person)
        db.flush()
        employee = Employee(person_id=person.id, employee_code="EMP-00012", job_title="Sale", status="active")
        db.add(employee)
        db.flush()
        db.add(EmployeeShiftSchedule(
            employee_id=employee.id,
            work_date=date(2026, 8, 17),
            starts_at=datetime(2026, 8, 17, 5, 0),
            ends_at=datetime(2026, 8, 17, 12, 0),
        ))
        db.commit()

        result = import_employee_shifts(db, {
            "sourceName": "lich.xlsx",
            "rows": [{
                "employeeName": "TRAN NGUYEN KHAI HOAN",
                "position": "Lễ tân",
                "shifts": [
                    {"workDate": "2026-08-17", "startTime": "05:00", "endTime": "12:00"},
                    {"workDate": "2026-08-17", "startTime": "17:00", "endTime": "22:00"},
                    {"workDate": "2026-08-18", "startTime": "14:00", "endTime": "22:00"},
                ],
            }],
        })

        rows = (
            db.query(EmployeeShiftSchedule)
            .filter(EmployeeShiftSchedule.employee_id == employee.id, EmployeeShiftSchedule.status == "active")
            .order_by(EmployeeShiftSchedule.starts_at.asc())
            .all()
        )
        assert result["created"] == 2
        assert len(result["skipped"]) == 1
        assert result["unmatched"] == []
        assert [(row.starts_at.hour, row.ends_at.hour) for row in rows] == [(5, 12), (17, 22), (14, 22)]
    finally:
        db.close()


def test_replace_employee_shifts_week_rewrites_selected_week(tmp_path):
    from server.models import Employee, EmployeeShiftSchedule, Person
    from server.services.operations_service import list_employee_shifts_week, replace_employee_shifts_week

    db = make_session(tmp_path)
    try:
        person = Person(display_name="Weekly Shift", phone="0900000013", status="active")
        db.add(person)
        db.flush()
        employee = Employee(person_id=person.id, employee_code="EMP-00013", job_title="Coach", status="active")
        db.add(employee)
        db.flush()
        db.add(EmployeeShiftSchedule(
            employee_id=employee.id,
            work_date=date(2026, 8, 17),
            starts_at=datetime(2026, 8, 17, 8, 0),
            ends_at=datetime(2026, 8, 17, 12, 0),
        ))
        db.commit()

        result = replace_employee_shifts_week(db, {
            "weekStart": "2026-08-17",
            "rows": [{
                "employeeId": employee.id,
                "employeeName": "Weekly Shift",
                "shifts": [
                    {"workDate": "2026-08-18", "startTime": "09:00", "endTime": "13:00"},
                    {"workDate": "2026-08-19", "startTime": "15:00", "endTime": "19:00"},
                ],
            }],
        })
        week = list_employee_shifts_week(db, "2026-08-17", "2026-08-23")

        assert result["created"] == 2
        assert result["deleted"] == 1
        assert [(row["workDate"], row["startTime"], row["endTime"]) for row in week["items"]] == [
            ("2026-08-18", "09:00", "13:00"),
            ("2026-08-19", "15:00", "19:00"),
        ]
    finally:
        db.close()


def test_employee_shift_report_marks_late_and_not_checked(tmp_path):
    from server.models import AttendanceSession, DahWebhookEvent, Employee, EmployeeShiftSchedule, Person
    from server.services.operations_service import employee_shift_report

    db = make_session(tmp_path)
    try:
        person = Person(display_name="Report Shift", phone="0900000014", status="active")
        db.add(person)
        db.flush()
        employee = Employee(person_id=person.id, employee_code="EMP-00014", job_title="Sale", status="active")
        db.add(employee)
        db.flush()
        first = EmployeeShiftSchedule(
            employee_id=employee.id,
            work_date=date(2026, 8, 17),
            starts_at=datetime(2026, 8, 17, 8, 0),
            ends_at=datetime(2026, 8, 17, 12, 0),
        )
        second = EmployeeShiftSchedule(
            employee_id=employee.id,
            work_date=date(2026, 8, 17),
            starts_at=datetime(2026, 8, 17, 14, 0),
            ends_at=datetime(2026, 8, 17, 18, 0),
        )
        third = EmployeeShiftSchedule(
            employee_id=employee.id,
            work_date=date(2026, 8, 17),
            starts_at=datetime(2026, 8, 17, 19, 0),
            ends_at=datetime(2026, 8, 17, 22, 0),
        )
        db.add_all([first, second, third])
        db.flush()
        session = AttendanceSession(
            employee_id=employee.id,
            employee_shift_schedule_id=first.id,
            scheduled_start_at=first.starts_at,
            scheduled_end_at=first.ends_at,
            checked_in_at=datetime(2026, 8, 17, 8, 12),
            checked_out_at=datetime(2026, 8, 17, 12, 0),
            source="dah",
            status="closed",
        )
        early_session = AttendanceSession(
            employee_id=employee.id,
            employee_shift_schedule_id=third.id,
            scheduled_start_at=third.starts_at,
            scheduled_end_at=third.ends_at,
            checked_in_at=datetime(2026, 8, 17, 19, 0),
            checked_out_at=datetime(2026, 8, 17, 21, 30),
            source="dah",
            status="closed",
        )
        db.add_all([session, early_session])
        db.flush()
        db.add_all([
            DahWebhookEvent(
                event_key="report-event-1",
                operator="face",
                employee_id=employee.id,
                attendance_session_id=session.id,
                event_time=datetime(2026, 8, 17, 8, 12),
                verify_status=1,
                status="processed",
                action="employee_shift_sync",
            ),
            DahWebhookEvent(
                event_key="report-event-2",
                operator="face",
                employee_id=employee.id,
                event_time=datetime(2026, 8, 17, 15, 45),
                verify_status=1,
                status="received",
                action="VerifyPush",
            ),
        ])
        db.commit()

        report = employee_shift_report(db, range_type="date", day="2026-08-17")
        shifts = report["items"][0]["days"][0]["shifts"]

        assert report["lateGraceMinutes"] == 10
        assert [shift["status"] for shift in shifts] == ["late", "not_checked", "early_checkout"]
        assert [shift["checkinStatus"] for shift in shifts] == ["late", "not_checked", "on_time"]
        assert [shift["checkoutStatus"] for shift in shifts] == ["on_time", "not_checked", "early_checkout"]
        assert shifts[0]["lateMinutes"] == 12
        assert shifts[0]["events"][0]["attendanceSessionId"] == session.id
        assert [event["eventTime"] for event in report["rows"][0]["dahEvents"]] == [
            "2026-08-17T08:12:00",
            "2026-08-17T15:45:00",
        ]
        assert shifts[1]["statusLabel"] == "Chưa chấm công"
        assert shifts[1]["checkinStatusLabel"] == "Chưa chấm công"
        assert shifts[2]["checkinStatusLabel"] == "Check-in đúng giờ"
        assert shifts[2]["checkoutStatusLabel"] == "Check-out sớm"
        assert shifts[2]["earlyCheckoutMinutes"] == 30
    finally:
        db.close()


def test_employee_shift_report_normalizes_status_and_filters_flat_rows(tmp_path):
    from server.models import Employee, EmployeeShiftSchedule, Person
    from server.services.operations_service import employee_shift_report

    db = make_session(tmp_path)
    try:
        people = [
            Person(display_name="Absent Sale", phone="0900000111", status="active"),
            Person(display_name="Future Coach", phone="0900000112", status="active"),
        ]
        db.add_all(people)
        db.flush()
        employees = [
            Employee(person_id=people[0].id, employee_code="EMP-00111", job_title="Sale", status="active"),
            Employee(person_id=people[1].id, employee_code="EMP-00112", job_title="Coach", status="active"),
        ]
        db.add_all(employees)
        db.flush()
        db.add_all([
            EmployeeShiftSchedule(
                employee_id=employees[0].id,
                work_date=date(2020, 1, 6),
                starts_at=datetime(2020, 1, 6, 8, 0),
                ends_at=datetime(2020, 1, 6, 12, 0),
            ),
            EmployeeShiftSchedule(
                employee_id=employees[1].id,
                work_date=date(2099, 1, 5),
                starts_at=datetime(2099, 1, 5, 18, 0),
                ends_at=datetime(2099, 1, 5, 22, 0),
            ),
        ])
        db.commit()

        absent = employee_shift_report(
            db,
            range_type="date",
            day="2020-01-06",
            status="anomaly",
            title="Sale",
            q="emp-00111",
            page=1,
            page_size=10,
        )
        future = employee_shift_report(
            db,
            range_type="date",
            day="2099-01-05",
            status="upcoming",
            shift_kind="night",
        )

        assert absent["summary"]["absent"] == 1
        assert absent["summary"]["pendingReview"] == 1
        assert absent["rows"][0]["displayStatus"] == "absent"
        assert absent["rows"][0]["needsReview"] is True
        assert absent["pagination"]["total"] == 1
        assert absent["filters"]["titles"] == ["Sale"]
        assert future["summary"]["upcoming"] == 1
        assert future["rows"][0]["displayStatus"] == "upcoming"
        assert future["rows"][0]["shiftKind"] == "night"
        assert future["rows"][0]["needsReview"] is False
    finally:
        db.close()


def test_employee_shift_report_sorts_by_employee_and_planned_time(tmp_path):
    from server.models import Employee, EmployeeShiftSchedule, Person
    from server.services.operations_service import employee_shift_report

    db = make_session(tmp_path)
    try:
        people = [
            Person(display_name="Bao Sale", phone="0900000121", status="active"),
            Person(display_name="An Coach", phone="0900000122", status="active"),
            Person(display_name="Chi Reception", phone="0900000123", status="active"),
        ]
        db.add_all(people)
        db.flush()
        employees = [
            Employee(person_id=people[0].id, employee_code="EMP-00121", job_title="Sale", status="active"),
            Employee(person_id=people[1].id, employee_code="EMP-00122", job_title="Coach", status="active"),
            Employee(person_id=people[2].id, employee_code="EMP-00123", job_title="Reception", status="active"),
        ]
        db.add_all(employees)
        db.flush()
        db.add_all([
            EmployeeShiftSchedule(
                employee_id=employees[0].id,
                work_date=date(2026, 8, 17),
                starts_at=datetime(2026, 8, 17, 14, 0),
                ends_at=datetime(2026, 8, 17, 18, 0),
            ),
            EmployeeShiftSchedule(
                employee_id=employees[1].id,
                work_date=date(2026, 8, 17),
                starts_at=datetime(2026, 8, 17, 8, 0),
                ends_at=datetime(2026, 8, 17, 12, 0),
            ),
            EmployeeShiftSchedule(
                employee_id=employees[2].id,
                work_date=date(2026, 8, 17),
                starts_at=datetime(2026, 8, 17, 19, 0),
                ends_at=datetime(2026, 8, 17, 22, 0),
            ),
        ])
        db.commit()

        by_employee_desc = employee_shift_report(db, range_type="date", day="2026-08-17", sort="employee_desc")
        by_planned_asc = employee_shift_report(db, range_type="date", day="2026-08-17", sort="planned_asc")
        by_planned_desc = employee_shift_report(db, range_type="date", day="2026-08-17", sort="planned_desc")

        assert [row["employeeName"] for row in by_employee_desc["rows"]] == [
            "Chi Reception",
            "Bao Sale",
            "An Coach",
        ]
        assert [(row["employeeName"], row["startTime"]) for row in by_planned_asc["rows"]] == [
            ("An Coach", "08:00"),
            ("Bao Sale", "14:00"),
            ("Chi Reception", "19:00"),
        ]
        assert [(row["employeeName"], row["startTime"]) for row in by_planned_desc["rows"]] == [
            ("Chi Reception", "19:00"),
            ("Bao Sale", "14:00"),
            ("An Coach", "08:00"),
        ]
    finally:
        db.close()


def test_employee_shift_report_uses_approved_override(tmp_path):
    from server.models import AttendanceSession, Employee, EmployeeShiftSchedule, Person
    from server.services.operations_service import approve_employee_shift_override, employee_shift_report

    db = make_session(tmp_path)
    try:
        person = Person(display_name="Override Shift", phone="0900000015", status="active")
        db.add(person)
        db.flush()
        employee = Employee(person_id=person.id, employee_code="EMP-00015", job_title="Sale", status="active")
        db.add(employee)
        db.flush()
        schedule = EmployeeShiftSchedule(
            employee_id=employee.id,
            work_date=date(2026, 8, 17),
            starts_at=datetime(2026, 8, 17, 8, 0),
            ends_at=datetime(2026, 8, 17, 12, 0),
        )
        db.add(schedule)
        db.flush()
        db.add(AttendanceSession(
            employee_id=employee.id,
            employee_shift_schedule_id=schedule.id,
            scheduled_start_at=schedule.starts_at,
            scheduled_end_at=schedule.ends_at,
            checked_in_at=datetime(2026, 8, 17, 14, 0),
            checked_out_at=datetime(2026, 8, 17, 22, 0),
            source="dah",
            status="closed",
        ))
        db.commit()

        before = employee_shift_report(db, range_type="date", day="2026-08-17")
        assert before["items"][0]["days"][0]["shifts"][0]["checkinStatus"] == "late"

        approve_employee_shift_override(db, schedule.id, {
            "workDate": "2026-08-17",
            "startTime": "14:00",
            "endTime": "22:00",
            "reason": "Đổi ca với team sale",
        })
        after = employee_shift_report(db, range_type="date", day="2026-08-17")
        shift = after["items"][0]["days"][0]["shifts"][0]

        assert shift["hasOverride"] is True
        assert shift["startTime"] == "14:00"
        assert shift["endTime"] == "22:00"
        assert shift["originalStartTime"] == "08:00"
        assert shift["checkinStatus"] == "on_time"
        assert shift["checkoutStatus"] == "on_time"
        assert shift["overrideReason"] == "Đổi ca với team sale"
    finally:
        db.close()


def test_preview_employee_shift_excel_reads_week_grid():
    from server.services.operations_service import preview_employee_shift_excel

    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "LỊCH LÀM - AM FITNESS & YOGA (17/08/2026 - 23/08/2026)"
    sheet["A2"] = "HỌ VÀ TÊN"
    sheet["B2"] = "VỊ TRÍ"
    for index, label in enumerate(["T2", "T3", "T4", "T5", "T6", "T7", "CN"], start=3):
        sheet.cell(row=2, column=index).value = label
        sheet.cell(row=3, column=index).value = 14 + index
    sheet["A4"] = "TRẦN NGUYỄN KHẢI HOÀN"
    sheet["B4"] = "LỄ TÂN+SALE"
    sheet["C4"] = "5H-12H\n17H-22H"
    sheet["D4"] = "OFF"
    buffer = BytesIO()
    workbook.save(buffer)

    preview = preview_employee_shift_excel(buffer.getvalue(), "lich.xlsx")

    assert preview["weekStart"] == "2026-08-17"
    assert preview["days"][0] == {"label": "T2", "workDate": "2026-08-17"}
    assert preview["rows"][0]["employeeName"] == "TRẦN NGUYỄN KHẢI HOÀN"
    assert preview["rows"][0]["position"] == "LỄ TÂN+SALE"
    assert preview["rows"][0]["cells"][:2] == ["5H-12H\n17H-22H", "OFF"]


def test_recent_checkins_filters_by_day_and_paginates(tmp_path):
    from server.models import AttendanceSession, Customer, Person
    from server.services.operations_service import recent_checkins

    db = make_session(tmp_path)
    try:
        person = Person(display_name="Checkin Member", phone="0900000020", status="active")
        db.add(person)
        db.flush()
        member = Customer(person_id=person.id, customer_code="CUS0000020", status="active")
        db.add(member)
        db.flush()
        db.add_all([
            AttendanceSession(
                customer_id=member.id,
                checked_in_at=datetime(2026, 8, 12, 9, 0),
                source="dah",
                status="open",
            ),
            AttendanceSession(
                customer_id=member.id,
                checked_in_at=datetime(2026, 8, 12, 8, 0),
                checked_out_at=datetime(2026, 8, 12, 10, 0),
                source="dah",
                status="closed",
            ),
            AttendanceSession(
                customer_id=member.id,
                checked_in_at=datetime(2026, 8, 11, 9, 0),
                source="dah",
                status="closed",
            ),
            AttendanceSession(
                customer_id=member.id,
                checked_in_at=datetime(2026, 8, 11, 22, 0),
                source="dah",
                status="open",
            ),
            AttendanceSession(
                customer_id=member.id,
                checked_in_at=datetime(2026, 8, 10, 22, 0),
                source="dah",
                status="open",
            ),
        ])
        db.commit()

        first_page = recent_checkins(db, day="2026-08-12", page=1, page_size=2)
        second_page = recent_checkins(db, day="2026-08-12", page=2, page_size=2)

        assert first_page["date"] == "2026-08-12"
        assert first_page["activeCount"] == 2
        assert first_page["pagination"]["total"] == 3
        assert [row["status"] for row in first_page["items"]] == ["open", "open"]
        assert second_page["items"][0]["status"] == "closed"
        assert recent_checkins(db, day="2026-08-11", page=1, page_size=20)["pagination"]["total"] == 3
    finally:
        db.close()


def test_recent_checkins_can_filter_members_and_employees_separately(tmp_path):
    from server.models import AttendanceSession, Customer, Employee, Person
    from server.services.operations_service import recent_checkins

    db = make_session(tmp_path)
    try:
        member_person = Person(display_name="Member Checkin", phone="0900000030", status="active")
        employee_person = Person(display_name="Employee Checkin", phone="0900000031", status="active")
        db.add_all([member_person, employee_person])
        db.flush()
        member = Customer(person_id=member_person.id, customer_code="CUS0000030", status="active")
        employee = Employee(person_id=employee_person.id, employee_code="EMP-00030", job_title="Sale", status="active")
        db.add_all([member, employee])
        db.flush()
        db.add_all([
            AttendanceSession(
                customer_id=member.id,
                checked_in_at=datetime(2026, 8, 12, 9, 0),
                source="dah",
                status="open",
            ),
            AttendanceSession(
                employee_id=employee.id,
                checked_in_at=datetime(2026, 8, 12, 9, 5),
                source="dah",
                status="open",
            ),
        ])
        db.commit()

        members = recent_checkins(db, day="2026-08-12", person_type="member", page=1, page_size=20)
        employees = recent_checkins(db, day="2026-08-12", person_type="employee", page=1, page_size=20)

        assert members["activeCount"] == 1
        assert employees["activeCount"] == 1
        assert [row["personType"] for row in members["items"]] == ["member"]
        assert [row["personType"] for row in employees["items"]] == ["employee"]
        assert members["pagination"]["total"] == 1
        assert employees["pagination"]["total"] == 1
    finally:
        db.close()


def test_recent_checkins_flags_members_with_membership_warnings(tmp_path):
    from server.models import AttendanceSession, Customer, Membership, Person, ServicePackage
    from server.services.operations_service import recent_checkins

    db = make_session(tmp_path)
    try:
        person = Person(display_name="Pending Checkin", phone="0900000021", status="active")
        db.add(person)
        db.flush()
        member = Customer(person_id=person.id, customer_code="CUS0000021", status="lead")
        plan = ServicePackage(
            code="FIT-WARN",
            name="Fitness Warning",
            category="Fitness",
            duration_days=30,
            price=100000,
            is_pt=False,
            is_active=True,
        )
        db.add_all([member, plan])
        db.flush()
        membership = Membership(
            customer_id=member.id,
            package_id=plan.id,
            code="MS-WARN",
            registered_at=date(2026, 8, 1),
            starts_at=date(2026, 8, 20),
            expires_at=date(2026, 9, 19),
            activated_at=None,
            status="pending",
        )
        checkin = AttendanceSession(
            customer_id=member.id,
            checked_in_at=datetime(2026, 8, 12, 9, 0),
            source="dah",
            status="open",
        )
        db.add_all([membership, checkin])
        db.commit()

        data = recent_checkins(db, day="2026-08-12", page=1, page_size=20)

        assert data["items"][0]["memberAccessWarning"] == "Gói đang chờ kích hoạt."
    finally:
        db.close()


def test_coach_can_only_update_assigned_pt_operational_fields(tmp_path):
    import pytest
    from fastapi import HTTPException
    from server.models import Customer, Employee, Person, PtEnrollment, PtEnrollmentCoach, User
    from server.services.operations_service import update_pt

    db = make_session(tmp_path)
    try:
        member_person = Person(display_name="PT Member", phone="0900000040")
        assigned_person = Person(display_name="Assigned Coach", phone="0900000041")
        other_person = Person(display_name="Other Coach", phone="0900000042")
        db.add_all([member_person, assigned_person, other_person])
        db.flush()
        member = Customer(person_id=member_person.id, customer_code="CUS-PT-AUTH", status="active")
        assigned = Employee(person_id=assigned_person.id, employee_code="EMP-PT-ASSIGNED", job_title="Coach", status="active")
        other = Employee(person_id=other_person.id, employee_code="EMP-PT-OTHER", job_title="Coach", status="active")
        db.add_all([member, assigned, other])
        db.flush()
        enrollment = PtEnrollment(customer_id=member.id, coach_id=assigned.id, total_sessions=12, remaining_sessions=12, status="active")
        db.add(enrollment)
        db.flush()
        enrollment.coach_assignments = [PtEnrollmentCoach(coach_id=assigned.id)]
        assigned_user = User(username="assigned-coach", display_name="Assigned Coach", password_hash="x", role="coach", employee_id=assigned.id)
        other_user = User(username="other-coach", display_name="Other Coach", password_hash="x", role="coach", employee_id=other.id)
        db.add_all([assigned_user, other_user])
        db.commit()

        updated = update_pt(db, enrollment.id, {"remainingSessions": 11}, assigned_user)
        assert updated["remainingSessions"] == 11
        with pytest.raises(HTTPException) as forbidden_owner:
            update_pt(db, enrollment.id, {"remainingSessions": 10}, other_user)
        assert forbidden_owner.value.status_code == 403
        with pytest.raises(HTTPException) as forbidden_field:
            update_pt(db, enrollment.id, {"totalSessions": 99}, assigned_user)
        assert forbidden_field.value.status_code == 403
    finally:
        db.close()


def test_pt_finance_tracks_debt_separately_from_membership_debt(tmp_path):
    from server.models import Customer, Payment, Person, PtDebtInstallment
    from server.services.dashboard_service import reports
    from server.services.operations_service import create_pt

    db = make_session(tmp_path)
    try:
        person = Person(display_name="BT Child", phone="0900000050", status="active")
        db.add(person)
        db.flush()
        member = Customer(person_id=person.id, customer_code="CUS-PT-DEBT", status="active")
        db.add(member)
        db.commit()

        data = create_pt(db, member.id, {
            "type": "1:1",
            "packageName": "BT Kids 12 buổi",
            "startsAt": "2026-08-20",
            "totalSessions": 12,
            "finalPrice": 1_000_000,
            "paidAmount": 400_000,
            "paidAt": "2026-08-20",
            "paymentMethod": "cash",
            "debtInstallments": [
                {"amount": 300_000, "dueDate": "2026-09-01"},
                {"amount": 300_000, "dueDate": "2026-10-01"},
            ],
        })

        assert data["finalPrice"] == 1_000_000
        assert data["packageName"] == "BT Kids 12 buổi"
        assert data["paidAmount"] == 400_000
        assert data["debtAmount"] == 600_000
        assert data["nextDebtDueDate"] == "2026-09-01"
        assert [row["amount"] for row in data["debtInstallments"]] == [300_000, 300_000]
        assert [row["paidAmount"] for row in data["debtInstallments"]] == [0, 0]
        assert [row["status"] for row in data["debtInstallments"]] == ["pending", "pending"]

        payment = db.query(Payment).one()
        assert payment.customer_id == member.id
        assert payment.membership_id is None
        assert payment.pt_enrollment_id == data["id"]
        assert payment.amount == 400_000
        assert payment.channel == "pt"
        assert payment.shift_date.isoformat() == "2026-08-20"

        stored_installments = db.query(PtDebtInstallment).order_by(PtDebtInstallment.due_date.asc()).all()
        assert [(row.amount, row.paid_amount, row.status) for row in stored_installments] == [
            (300_000, 0, "pending"),
            (300_000, 0, "pending"),
        ]

        report = reports(db, "2026-08-20", "2026-08-20")
        assert report["summary"]["revenue"] == 400_000
        assert report["summary"]["membershipRevenue"] == 0
        assert report["summary"]["ptRevenue"] == 400_000
        assert report["summary"]["debt"] == 0
        assert report["debts"] == []
        pt_type = next(row for row in report["revenueByType"] if row["type"] == "pt")
        assert pt_type["amount"] == 400_000
        assert pt_type["payments"] == 1
        assert report["daily"][0]["ptAmount"] == 400_000
        assert report["daily"][0]["membershipAmount"] == 0
        assert report["revenueItems"][0]["type"] == "pt"
        assert report["revenueItems"][0]["ptEnrollmentId"] == data["id"]
    finally:
        db.close()


def test_pt_finance_validates_paid_amount_installments_and_transfer_account(tmp_path):
    import pytest
    from fastapi import HTTPException
    from server.models import Customer, Person
    from server.services.operations_service import create_pt

    db = make_session(tmp_path)
    try:
        person = Person(display_name="BT Validation", phone="0900000051", status="active")
        db.add(person)
        db.flush()
        member = Customer(person_id=person.id, customer_code="CUS-PT-VALID", status="active")
        db.add(member)
        db.commit()

        with pytest.raises(HTTPException) as overpaid:
            create_pt(db, member.id, {
                "finalPrice": 500_000,
                "paidAmount": 600_000,
                "debtInstallments": [],
            })
        assert overpaid.value.status_code == 422

        with pytest.raises(HTTPException) as bad_installments:
            create_pt(db, member.id, {
                "finalPrice": 1_000_000,
                "paidAmount": 400_000,
                "debtInstallments": [{"amount": 500_000, "dueDate": "2026-09-01"}],
            })
        assert bad_installments.value.status_code == 422

        with pytest.raises(HTTPException) as missing_account:
            create_pt(db, member.id, {
                "finalPrice": 1_000_000,
                "paidAmount": 1_000_000,
                "paymentMethod": "bank_transfer",
                "debtInstallments": [],
            })
        assert missing_account.value.status_code == 422
    finally:
        db.close()


def test_pt_finance_update_collects_more_money_and_replaces_remaining_debt_plan(tmp_path):
    from server.models import Customer, Payment, Person, PtDebtInstallment
    from server.services.operations_service import create_pt, update_pt

    db = make_session(tmp_path)
    try:
        person = Person(display_name="BT Update", phone="0900000052", status="active")
        db.add(person)
        db.flush()
        member = Customer(person_id=person.id, customer_code="CUS-PT-UPD", status="active")
        db.add(member)
        db.commit()

        created = create_pt(db, member.id, {
            "type": "1:2",
            "packageName": "PT đôi 10 buổi",
            "startsAt": "2026-08-20",
            "totalSessions": 10,
            "finalPrice": 1_000_000,
            "paidAmount": 400_000,
            "paidAt": "2026-08-20",
            "debtInstallments": [
                {"amount": 300_000, "dueDate": "2026-09-01"},
                {"amount": 300_000, "dueDate": "2026-10-01"},
            ],
        })

        updated = update_pt(db, created["id"], {
            "packageName": "PT đôi 10 buổi - đã thu thêm",
            "finalPrice": 1_000_000,
            "paidAmount": 700_000,
            "paidAt": "2026-09-05",
            "paymentMethod": "cash",
            "debtInstallments": [{"amount": 300_000, "dueDate": "2026-10-01"}],
        })

        assert updated["paidAmount"] == 700_000
        assert updated["packageName"] == "PT đôi 10 buổi - đã thu thêm"
        assert updated["debtAmount"] == 300_000
        assert len(updated["debtInstallments"]) == 1
        assert updated["debtInstallments"][0]["amount"] == 300_000
        assert updated["debtInstallments"][0]["paidAmount"] == 0
        assert updated["debtInstallments"][0]["status"] == "pending"

        payments = db.query(Payment).order_by(Payment.id.asc()).all()
        assert [row.amount for row in payments] == [400_000, 300_000]
        assert [row.payment_no for row in payments] == [
            f"PTPAY-{created['id']:06d}-001",
            f"PTPAY-{created['id']:06d}-002",
        ]
        installments = db.query(PtDebtInstallment).order_by(PtDebtInstallment.due_date.asc()).all()
        assert [(row.amount, row.paid_amount, row.status) for row in installments] == [(300_000, 0, "pending")]
    finally:
        db.close()
